import openpyxl # type: ignore # False alert with Pylance
class ExcelSheet:
    def __init__(self, path, sheet_name=None):
        """
        Initialize an Excel sheet reader.

        This initializer loads an Excel workbook from the specified path and
        selects a worksheet for subsequent data access. If no sheet name is
        provided, the first worksheet in the workbook is selected by default.
        Basic sheet dimensions are stored for later use.

        Parameters
        ----------
        path : str
            Path to the Excel file to be loaded.
        sheet_name : str, optional
            Name of the worksheet to select. If ``None``, the first worksheet
            in the workbook is used.
        """
        self.workbook = openpyxl.load_workbook(filename=path, data_only=True)
        #if sheet_name is None and len(self.workbook.sheetnames) == 1:
        if sheet_name is None:
            sheet_name = self.workbook.sheetnames[0]
        #else:
        #    print("ERROR - ExcelSheet - please give a sheet name for file <"+path+">, found ones are",self.workbook.sheetnames,"")
        #    exit()
        self.sheet = self.workbook[sheet_name]
        self.max_col = min(200, self.sheet.max_column)
        self.max_row = self.sheet.max_row
        #print(self.max_col)
        
    #def _letter2number(self, letter):
    #    return "ABCDEFGHIJKLMNOPQRSTUVWXYZ".index(letter)
    #def _number2letter(self, number):
    #    print(number)
    #    return "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[number]
    def _number2letter(self, number):
        '''
        return the "number2-th letter of the alphabet
        
        :param number: int

        :return: string
        '''
        #print(number)
        def _toto(n):
            #print(n, len("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
            return "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[n]
        if number<26: return _toto(number)
        return _toto(number//26 - 1)+_toto(number%26)
    def get(self, i,j):
        return self.sheet[self._number2letter(j)+str(i+1)].value
    def set(self, i,j,v):
        self.sheet[self._number2letter(j)+str(i+1)] = v
    def save(self, path):
        self.workbook.save(filename=path)
    def get_nb_col(self):
        #return self.sheet.max_column
        return self.max_col
    def get_nb_row(self):
        #return self.sheet.max_row
        return self.max_row
    def sub_list(self, column_to_keep_no, condition=None, condition_column_no= None):
        """
        Extract a sub-list of values from a specified column, optionally filtered
        by a condition.

        The method iterates over the rows of the sheet and collects values from
        the specified column. If a condition is provided, values are included only
        for rows matching the condition, either across all columns or within a
        specific column.

        Parameters
        ----------
        column_to_keep_no : int
            Index of the column whose values should be extracted.
        condition : object, optional
            Value used to filter rows. If ``None``, all values from the specified
            column are returned.
        condition_column_no : int, optional
            Index of the column in which the condition is evaluated.
            If ``None`` and a condition is provided, the condition is searched for
            across all columns in each row.

        Returns
        -------
        list
            List of values extracted from the specified column that satisfy the
            given condition.
        """
        sub_l = []
        if condition==None:
            for i in range(self.get_nb_row()):
                sub_l.append(self.get(i,column_to_keep_no))
        elif condition_column_no == None:
            for i in range(self.get_nb_row()):
                for j in range(self.get_nb_col()):
                    if self.get(i,j)==condition :
                        sub_l.append(self.get(i,column_to_keep_no))
        else:
            for i in range(self.get_nb_row()):
                if self.get(i,condition_column_no)==condition :
                    sub_l.append(self.get(i,column_to_keep_no))
        return sub_l
